#Abspeicherungsprogramm für Excel von Jery:
#In diesem Programm wird veranschaulicht wie eine Partie in Excel abgespeichert wird.
#Für dies wird die Implementation "openpyxl" genutzt.
#Link zur Dokumentation: https://openpyxl.readthedocs.io/en/stable/#

import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.styles.borders import BORDER_THIN

# Dateityp: xlsx
path = 'Pfad der Datei'
Excel = xl.load_workbook(path)
tabelname = Excel['Name der Tabelle']


def setValue(row, column, tabel, value):
    tabel.cell(row=row, column=column).value = value


#Die GetValue Methode braucht man später für die Daten Auswertung.
def getValue(row, column, tabel):
    return tabel.cell(row=row, column=column).value


#Diese Methode ist optional. Sie dient nur für ein überschaubares Design.
def setStyle(row, column, tabel):
    tabel.cell(row=row, column=column).fill = PatternFill(fgColor="dadada", fill_type="solid")
    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='bfbfbf'),
        right=Side(border_style=BORDER_THIN, color='bfbfbf'),
        top=Side(border_style=BORDER_THIN, color='bfbfbf'),
        bottom=Side(border_style=BORDER_THIN, color='bfbfbf')
    )
    tabel.cell(row=row, column=column).border = thin_border


#Die gameID muss bei jedem Eintrag verändert werden.
#Der kleinste Wert ist 0 und man kann so viele Partien eintragen wie man will.
#Bei jedem zusätzlichen Eintrag muss man den Wert um 1 erhöhen.
#Es werden 12 Zeilen und 7 Spalten pro GameID bearbeitet.
gameID = 0

#Liste welche das Feld abspeichert.
field = []

#Da wird die Anzahl der Züge gespeichert.
steps = 0

#Startliste wird generiert.
for i in range(42):
    field.append("nn")

#Eintragen der Daten in Excel.
for i in range(6):
    for k in range(7):
        setValue(12 * gameID + i + 1, k + 1, tabelname, field[i * 7 + k])
        setStyle(12 * gameID + i + 1, k + 1, tabelname) #(optional)

setValue(gameID * 12 + 8, 1, tabelname, "Steps:")
setValue(gameID * 12 + 8, 2, tabelname, steps)
setValue(gameID * 12 + 9, 1, tabelname, "R Value:")
setValue(gameID * 12 + 10, 1, tabelname, "Y Value:")

#Excel Datei wird abgespeichert.
Excel.save(path)
