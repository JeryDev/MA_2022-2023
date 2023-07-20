# Es werden die Daten von dem Grundgerüst in ein neues Format umgewandelt

import os
import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.styles.borders import BORDER_THIN

sheet_name = "Spieler1 vs Spieler2"
player_1 = sheet_name.split(" ")[0]
player_2 = sheet_name.split(" ")[2]

# Werte aus Excel abrufen
def getValue(row: int, column: int, tabel):
    return tabel.cell(row=row, column=column).value


# Werte in Excel festlegen
def setValue(row: int, column: int, worksheet, value):
    worksheet.cell(row=row, column=column).value = value

def setStyle(row: int, column: int, worksheet):
    worksheet.cell(row=row, column=column).fill = PatternFill(
        fgColor="dadada", fill_type="solid"
    )
    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color="bfbfbf"),
        right=Side(border_style=BORDER_THIN, color="bfbfbf"),
        top=Side(border_style=BORDER_THIN, color="bfbfbf"),
        bottom=Side(border_style=BORDER_THIN, color="bfbfbf"),
    )
    worksheet.cell(row=row, column=column).border = thin_border

def checkWin(board: list):
    """
    Überprüft ob jemand gewonnen hat

    Args:
        board: Eine Liste, die das Spielbrett repräsentiert.

    Returns:
        yellow: Gelb hat gewonnen.
        red: Rot hat gewonnen.
        None: Niemand hat gewonnen.
    """
    won = None

    # Überprüfung der horizontalen Gewinnbedingung
    for row in range(6):
        for column in range(4):
            cell_index = column + row * 7
            if board[cell_index][0] != "n":
                if all(
                    board[cell_index + i][0] == board[cell_index][0]
                    for i in range(1, 4)
                ):
                    if board[cell_index][0] == "y":
                        return "yellow"
                    else:
                        return "red"

    # Überprüfung der vertikalen Gewinnbedingung
    for column in range(7):
        for row in range(3):
            cell_index = column + row * 7
            if board[cell_index][0] != "n":
                if all(
                    board[cell_index + i * 7][0] == board[cell_index][0]
                    for i in range(1, 4)
                ):
                    if board[cell_index][0] == "y":
                        return "yellow"
                    else:
                        return "red"

    # Überprüfung der Diagonalen (rechts unten nach links oben) Gewinnbedingung
    for row in range(3):
        for column in range(4):
            cell_index = column + row * 7
            if board[cell_index][0] != "n":
                if all(
                    board[cell_index + i * 8][0] == board[cell_index][0]
                    for i in range(1, 4)
                ):
                    if board[cell_index][0] == "y":
                        return "yellow"
                    else:
                        return "red"

    # Überprüfung der Diagonalen (links unten nach rechts oben) Gewinnbedingung
    for row in range(3):
        for column in range(4):
            cell_index = column + 3 + row * 7
            if board[cell_index][0] != "n":
                if all(
                    board[cell_index + i * 6][0] == board[cell_index][0]
                    for i in range(1, 4)
                ):
                    if board[cell_index][0] == "y":
                        return "yellow"
                    else:
                        return "red"

    return won

file_path_data = "Pfad zur Datendatei"
file_path_results = os.path.join(os.getcwd(), "results.xlsx")


if not os.path.isfile(file_path_results):
    workbook_results = xl.Workbook()
    worksheet_results = workbook_results.active
    worksheet_results.title = "Stats"
    workbook_results.save(file_path_results)

workbook_data = xl.load_workbook(file_path_data)
workbook_results = xl.load_workbook(file_path_results)

try:
    worksheet_results = workbook_results[player_1]

except KeyError:
    workbook_results.create_sheet(player_1)
    workbook_results.save(file_path_results)
    workbook_results = xl.load_workbook(file_path_results)
    worksheet_player_1 = workbook_results[player_1]
    setValue(1, 1, worksheet_player_1, "Date")
    setStyle(1, 1, worksheet_player_1)
    setValue(1, 2, worksheet_player_1, "Avg Time")
    setStyle(1, 2, worksheet_player_1)
    setValue(1, 3, worksheet_player_1, "Accuracy")
    setStyle(1, 3, worksheet_player_1)
    setValue(1, 4, worksheet_player_1, "Blunder")
    setStyle(1, 4, worksheet_player_1)
    setValue(1, 5, worksheet_player_1, "Opponent")
    setStyle(1, 5, worksheet_player_1)
    setValue(1, 6, worksheet_player_1, "Status")
    setStyle(1, 6, worksheet_player_1)
    setValue(1, 9, worksheet_player_1, "ID:")
    setStyle(1, 9, worksheet_player_1)
    setValue(1, 10, worksheet_player_1, 1)
    setStyle(1, 10, worksheet_player_1)
    workbook_results.save(file_path_results)
    workbook_results = xl.load_workbook(file_path_results)

try:
    worksheet_results = workbook_results[player_2]

except KeyError:
    workbook_results.create_sheet(player_2)
    workbook_results.save(file_path_results)
    workbook_results = xl.load_workbook(file_path_results)
    worksheet_player_2 = workbook_results[player_2]
    setValue(1, 1, worksheet_player_2, "Date")
    setStyle(1, 1, worksheet_player_2)
    setValue(1, 2, worksheet_player_2, "Avg Time")
    setStyle(1, 2, worksheet_player_2)
    setValue(1, 3, worksheet_player_2, "Accuracy")
    setStyle(1, 3, worksheet_player_2)
    setValue(1, 4, worksheet_player_2, "Blunder")
    setStyle(1, 4, worksheet_player_2)
    setValue(1, 5, worksheet_player_2, "Opponent")
    setStyle(1, 5, worksheet_player_2)
    setValue(1, 6, worksheet_player_2, "Status")
    setStyle(1, 6, worksheet_player_2)    
    setValue(1, 9, worksheet_player_2, "ID:")
    setStyle(1, 9, worksheet_player_2)
    setValue(1, 10, worksheet_player_2, 1)
    setStyle(1, 10, worksheet_player_2)
    workbook_results.save(file_path_results)
    workbook_results = xl.load_workbook(file_path_results)

worksheet_data = workbook_data[sheet_name]
worksheet_player_1 = workbook_results[player_1]
worksheet_player_2 = workbook_results[player_2]

gameID = int(getValue(1, 10, worksheet_data))

for i in range(gameID):
    board_from_excel = []
    for row in range(6):
        for column in range(7):
            board_from_excel.append(str(getValue((i * 14) + row + 1, column + 1, worksheet_data)))
    winner = checkWin(board_from_excel)
    
    if str(getValue(i * 14 + 9, 2, worksheet_data)) == player_1:
        id = int(getValue(1, 10, worksheet_player_1)) + 1
        setValue(1, 10, worksheet_player_1, id)
        date = getValue(i * 14 + 12, 2, worksheet_data)
        avg_time = getValue(i * 14 + 9, 3, worksheet_data)
        accuracy = getValue(i * 14 + 9, 4, worksheet_data)
        blunder = getValue(i * 14 + 9, 5, worksheet_data)
        setValue(id, 1, worksheet_player_1, date)
        setValue(id, 2, worksheet_player_1, avg_time)
        setValue(id, 3, worksheet_player_1, accuracy)
        setValue(id, 4, worksheet_player_1, blunder)
        setValue(id, 5, worksheet_player_1, player_2)
        if winner == "yellow":
            setValue(id, 6, worksheet_player_1, "Win")
        elif winner == "red":
            setValue(id, 6, worksheet_player_1, "Lose")
        else:
            setValue(id, 6, worksheet_player_1, "Draw")
    
    else:
        id = int(getValue(1, 10, worksheet_player_1)) + 1
        setValue(1, 10, worksheet_player_1, id)
        date = getValue(i * 14 + 12, 2, worksheet_data)
        avg_time = getValue(i * 14 + 10, 3, worksheet_data)
        accuracy = getValue(i * 14 + 10, 4, worksheet_data)
        blunder = getValue(i * 14 + 10, 5, worksheet_data)
        setValue(id, 1, worksheet_player_1, date)
        setValue(id, 2, worksheet_player_1, avg_time)
        setValue(id, 3, worksheet_player_1, accuracy)
        setValue(id, 4, worksheet_player_1, blunder)
        setValue(id, 5, worksheet_player_1, player_2)
        if winner == "red":
            setValue(id, 6, worksheet_player_1, "Win")
        elif winner == "yellow":
            setValue(id, 6, worksheet_player_1, "Lose")
        else:
            setValue(id, 6, worksheet_player_2, "Draw")
    
    if str(getValue(i * 14 + 9, 2, worksheet_data)) == player_2:
        id = int(getValue(1, 10, worksheet_player_2)) + 1
        setValue(1, 10, worksheet_player_2, id)
        date = getValue(i * 14 + 12, 2, worksheet_data)
        avg_time = getValue(i * 14 + 9, 3, worksheet_data)
        accuracy = getValue(i * 14 + 9, 4, worksheet_data)
        blunder = getValue(i * 14 + 9, 5, worksheet_data)
        setValue(id, 1, worksheet_player_2, date)
        setValue(id, 2, worksheet_player_2, avg_time)
        setValue(id, 3, worksheet_player_2, accuracy)
        setValue(id, 4, worksheet_player_2, blunder)
        setValue(id, 5, worksheet_player_2, player_1)
        if winner == "yellow":
            setValue(id, 6, worksheet_player_2, "Win")
        elif winner == "red":
            setValue(id, 6, worksheet_player_2, "Lose")
        else:
            setValue(id, 6, worksheet_player_2, "Draw")
    
    else:
        id = int(getValue(1, 10, worksheet_player_2)) + 1
        setValue(1, 10, worksheet_player_2, id)
        date = getValue(i * 14 + 12, 2, worksheet_data)
        avg_time = getValue(i * 14 + 10, 3, worksheet_data)
        accuracy = getValue(i * 14 + 10, 4, worksheet_data)
        blunder = getValue(i * 14 + 10, 5, worksheet_data)
        setValue(id, 1, worksheet_player_2, date)
        setValue(id, 2, worksheet_player_2, avg_time)
        setValue(id, 3, worksheet_player_2, accuracy)
        setValue(id, 4, worksheet_player_2, blunder)
        setValue(id, 5, worksheet_player_2, player_1)
        if winner == "red":
            setValue(id, 6, worksheet_player_2, "Win")
        elif winner == "yellow":
            setValue(id, 6, worksheet_player_2, "Lose")
        else:
            setValue(id, 6, worksheet_player_2, "Draw")


workbook_results.save(file_path_results)
print("Done")
