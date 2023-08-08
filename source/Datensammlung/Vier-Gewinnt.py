"""
Vier-Gewinnt Python-Programm (pygame) von Jery (getestet mit: Python 3.11)

In diesem Programm treten zwei verschiedene Spieler gegeneinander an und spielen Vier-Gewinnt.
- Es wird ein Ordner, der den Namen "Maturaarbeit_Jery" trägt, erstellt.
- In diesem Ordner sind folgende Dateien enthalten: config.json und data.xlsx


In der config.json kann man die Grösse des Fensters anpassen.
Standard:
{
 "size": 800
}


In einer Excel Datei werden die folgenden Daten abgespeichert:
- Reihenfolge der Züge (inkl. Farben)
- Anzahl Züge insgesamt
- Das Datum, an welchem gespielt wurde
- Namen der Spieler (zu den Farben zugeordnet)
- Durchschnittliche Zeit für einen Zug (beide Spieler)
- Qualität eines Spielers in Prozent (beide Spieler)
- Anzahl Patzer (Blunder), welche in einem Spiel gemacht wurden (beide Spieler)


Anleitung:
1) Python 3.11 herunterladen: https://www.python.org/downloads/
2) In der Eingabeaufforderung (Windows: cmd / MacOS: terminal) diesen Befehl testen: python --version oder python3 --version
3) Wenn der vorherige Schritt erfolgreich war, diese Datei herunterladen
4) Lege die Vier-Gewinnt.py in einen Ordner oder auf den Desktop, damit du die Datei wiederfindest
   → Dieser Ort wird dann zum Speicherort von allen Daten.
   → Sobald ein Spiel gespielt wurde, dürfen nur alle Dateien zusammen verschoben werden!
5) Alle nötigen Pakete mit diesem pip herunterladen: pip install package oder pip3 install package
6) In der Eingabeaufforderung (Windows: cmd / MacOS: terminal) den Code ausführen: python Vier-Gewinnt.py oder python3 Vier-Gewinnt.py
7) Titelformat: Spieler1 vs Spieler2
8) Überprüfe, ob sich die Datei in deinem erstellten Ordner (Schritt vier) oder auf dem Desktop befindet.


Mögliche Schwierigkeiten / Probleme, welche während der Anleitung auftreten können:
- Bei der Installation muss unter "Customize installation" überprüft werden, dass Auswahlfeld pip ausgewählt ist.
- Beim fünften Schritt muss der Pfad der Datei miteinbezogen werden: 
  1) Pfad der Datei (Vier-Gewinnt.py) herausfinden
  2) In der Eingabeaufforderung (Windows: cmd / MacOS: terminal) zuerst mit dem Befehl: "cd Pfad" zur Datei navigieren
  3) Dann die Datei wie in Schritt fünf ausführen
- Pygame lässt sich möglicherweise nicht downloaden. Hier ist die genaue Anleitung: https://www.pygame.org/news
  

Pakete:
→ https://github.com/JeryDev/MA_2022-2023/blob/main/source/Datensammlung/modules.txt
"""


import pygame as p
import os
import json
import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.styles.borders import BORDER_THIN
from datetime import date
import time

# Initialisierungsfunktion
p.init()

# Pfade für die nötigen Dateien
folder_path = os.path.join(os.getcwd(), "Maturaarbeit_Jery")
file_path_xlsx = os.path.join(folder_path, "data.xlsx")
file_path_config = os.path.join(folder_path, "config.json")

# Titel der Mappe -> Format: Spieler1 vs Spieler2
sheet_name = str(input("Geben Sie einen Titel ein: "))

# Name des Gelben Spieler -> Spieler1 oder Spieler2
yellowPlayer = str(input("Geben Sie den gelben Spieler an: "))
if sheet_name.split(" ")[0] == yellowPlayer:
    redPlayer = sheet_name.split(" ")[2]
else:
    redPlayer = sheet_name.split(" ")[0]


# Werte aus Excel abrufen
def getValue(row: int, column: int, tabel):
    return tabel.cell(row=row, column=column).value


# Werte in Excel festlegen
def setValue(row: int, column: int, worksheet, value):
    worksheet.cell(row=row, column=column).value = value


# Überschaubares Design für einzelne Spalten setzen
def setStyle(row: int, column: int, worksheet):
    worksheet.cell(row=row, column=column).fill = PatternFill(
        fgColor="dadada", fill_type="solid"
    )
    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color="bfbfbf"),
        right=Side(border_style=BORDER_THIN, color="bfbfbf"),
        top=Side(border_style=BORDER_THIN, color="bfbfbf"),
        bottom=Side(border_style=BORDER_THIN, color="bfbfbf"),
    )
    worksheet.cell(row=row, column=column).border = thin_border


# Überprüfen ob es den Ordner bereits gibt
# Wenn nicht dann einen neuen Ordner anlegen
if not os.path.isdir(folder_path):
    os.mkdir(folder_path)

# Überprüfen ob es die Excel Datei bereits gibt
# Wenn nicht dann einen neuen Datei anlegen mit Standardwert
if not os.path.isfile(file_path_xlsx):
    workbook = xl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Stats"
    workbook.save(file_path_xlsx)

workbook = xl.load_workbook(file_path_xlsx)


# Überprüfen ob es die Konfigurationsdatei bereits gibt
# Wenn nicht dann einen neuen Datei anlegen mit Standardwert
if not os.path.isfile(file_path_config):
    data = {"size": 800}
    with open(file_path_config, "w") as file:
        json.dump(data, file, indent=1)


# Die Konfigurationsdatei öffnen
f = open(file_path_config)

# Konfigurationsdatei in ein Dictionary umwandeln
data = json.load(f)

# Den "Size" Parameter abrufen
size = data["size"]
f.close()

# Grösse von einem Feld bestimmen
sizeOne = int(size / 7)

width = size
height = size

# Hier werden die Grössen eingetragen.
screen = p.display.set_mode((width, height))

# Fensterbeschriftung
p.display.set_caption("Vier-Gewinnt")

# Liste, welche die besetzten Felder speichert
occupied = []

# Liste, welche das aktuelle Feld speichert
board = []

# Leeres Feld wird generiert
for n in range(42):
    board.append("nnn")

# Aktueller Spielzug
move = 0


def isValidDirection(next_index: int, end: int, direction: str):
    """
    Überprüft, ob ein Index in der gegebenen Richtung gültig ist, wenn er als
    nächstes besucht wird.

    Args:
        next_index: Der nächste Index, der überprüft werden soll.
        end: Der Index, von dem aus gestartet wird.
        direction: Die Richtung, in der überprüft werden soll.

    Returns:
        True, wenn der nächste Index in der gegebenen Richtung gültig ist, andernfalls False.
    """
    if direction == "vertical":
        if getRow(end) + 1 == getRow(next_index):
            # Der nächste Index ist eine Zeile darunter
            return True
        elif getRow(end) - 1 == getRow(next_index):
            # Der nächste Index ist eine Zeile darüber
            return True
    elif direction == "horizontal":
        if getColumn(end) + 1 == getColumn(next_index):
            # Der nächste Index ist eine Spalte rechts
            return True
        elif getColumn(end) - 1 == getColumn(next_index):
            # Der nächste Index ist eine Spalte links
            return True
    elif direction == "diagonal":
        if getRow(end) + 1 == getRow(next_index) and getColumn(end) + 1 == getColumn(
            next_index
        ):
            # Der nächste Index ist diagonal unten rechts
            return True
        elif getRow(end) - 1 == getRow(next_index) and getColumn(end) - 1 == getColumn(
            next_index
        ):
            # Der nächste Index ist diagonal oben links
            return True
        elif getRow(end) + 1 == getRow(next_index) and getColumn(end) - 1 == getColumn(
            next_index
        ):
            # Der nächste Index ist diagonal unten links
            return True
        elif getRow(end) - 1 == getRow(next_index) and getColumn(end) + 1 == getColumn(
            next_index
        ):
            # Der nächste Index ist diagonal oben rechts
            return True
    # Der nächste Index ist in keiner der erlaubten Richtungen gültig.
    return False


def getRow(index: int):
    """
    Gibt die Zeile zurück, in der sich der gegebene Index befindet.

    Args:
        index: Der Index, dessen Zeile zurückgegeben werden soll.

    Returns:
        Die Zeile, in der sich der gegebene Index befindet.
    """
    return index // 7 + 1


def getColumn(index: int):
    """
    Gibt die Spalte zurück, in der sich der gegebene Index befindet.

    Args:
        index: Der Index, dessen Spalte zurückgegeben werden soll.

    Returns:
        Die Spalte, in der sich der gegebene Index befindet.
    """
    return index % 7 + 1


def getColor(index: int, board: list):
    """
    Gibt die Farbe des Steins am gegebenen Index auf dem gegebenen Spielbrett zurück.

    Args:
        index: Der Index, an dem sich der Stein befindet.
        board: Eine Liste, die das Spielbrett repräsentiert.

    Returns:
        Die Farbe des Steins am gegebenen Index auf dem Spielbrett.
        Gibt "n" zurück, wenn kein Stein an diesem Index vorhanden ist.
    """
    if index < len(
        board
    ):  # Überprüfen, ob der gegebene Index im Bereich des Spielbretts liegt
        return board[index][0]
    else:
        return "n"

def getUnoccupiedLocations(board: list):
    """
    Gibt eine Liste von Indexe zurück, die auf freie Plätze auf dem Spielbrett verweisen.

    Args:
        board: Eine Liste, die das Spielbrett repräsentiert.

    Returns:
        Eine Liste von Indexe, die auf freie Plätze auf dem Spielbrett verweisen.
    """
    free_spaces = []
    for i in range(7):
        for k in range(6):
            # Berechnung des Index des Elements in der aktuellen Spalte und Zeile.
            v = i + ((5 - k) * 7)
            if board[v][0] == "n":
                free_spaces.append(v)
                break
    return free_spaces


def isMiddle(index: int):
    """
    Überprüft, ob der gegebene Index in der Mitte auf dem Spielbrett ist.

    Args:
        index: Der Index, der überprüft werden soll.

    Returns:
        True, wenn der gegebene Index in der Mitte auf dem Spielbrett ist.
        False, wenn der gegebene Index nicht in der Mitte auf dem Spielbrett ist.
    """
    return index in (3, 10, 17, 24, 31, 38)


DIRECTION_LOOKUP = {
    -8: (-1, -1, "diagonal"),
    -7: (-1, 0, "vertical"),
    -6: (-1, 1, "diagonal"),
    -1: (0, -1, "horizontal"),
    1: (0, 1, "horizontal"),
    6: (1, -1, "diagonal"),
    7: (1, 0, "vertical"),
    8: (1, 1, "diagonal"),
}


def getDirection(start: int, end: int):
    """
    Bestimmt die Richtung basierend auf dem Start- und Endindex.

    Args:
        start: Der Startindex.
        end: Der Endindex.

    Returns:
        Der nächste Index in der berechneten Richtung, wenn die Richtung gültig ist.
        False, wenn die Richtung ungültig ist oder der nächste Index außerhalb des Spielbretts liegt.
    """
    diff = end - start  # Die Differenz zwischen Start- und Endindex wird berechnet.
    # Wenn die Differenz in der DIRECTION_LOOKUP-Liste enthalten ist, gibt es eine gültige Richtung.
    if diff in DIRECTION_LOOKUP:
        # Die Zeilen- und Spaltenunterschiede sowie die Richtung werden aus der DIRECTION_LOOKUP-Liste abgerufen.
        row_diff, col_diff, direction = DIRECTION_LOOKUP[diff]
        # Der nächste Index in der berechneten Richtung wird berechnet.
        next_index = end + row_diff * 7 + col_diff
        # Überprüfung, ob der nächste Index in der berechneten Richtung gültig ist und innerhalb des Spielbretts liegt.
        if isValidDirection(next_index, end, direction) and 0 <= next_index <= 41:
            return next_index  # Der nächste Index wird zurückgegeben.
    # Wenn die Richtung ungültig ist oder der nächste Index außerhalb des Spielbretts liegt, wird False zurückgegeben.
    return False


def checkWin(board: list):
    """
    Überprüft ob jemand gewonnen hat

    Args:
        board: Eine Liste, die das Spielbrett repräsentiert.

    Returns:
        yellow: Gelb hat gewonnen.
        red: Rot hat gewonnen.
        None: Niemand hat gewonnen.
    """
    won = None

    # Überprüfung der horizontalen Gewinnbedingung
    for row in range(6):
        for column in range(4):
            cell_index = column + row * 7
            if board[cell_index][0] != "n":
                if all(
                    board[cell_index + i][0] == board[cell_index][0]
                    for i in range(1, 4)
                ):
                    if board[cell_index][0] == "y":
                        return "yellow"
                    else:
                        return "red"

    # Überprüfung der vertikalen Gewinnbedingung
    for column in range(7):
        for row in range(3):
            cell_index = column + row * 7
            if board[cell_index][0] != "n":
                if all(
                    board[cell_index + i * 7][0] == board[cell_index][0]
                    for i in range(1, 4)
                ):
                    if board[cell_index][0] == "y":
                        return "yellow"
                    else:
                        return "red"

    # Überprüfung der Diagonalen (rechts unten nach links oben) Gewinnbedingung
    for row in range(3):
        for column in range(4):
            cell_index = column + row * 7
            if board[cell_index][0] != "n":
                if all(
                    board[cell_index + i * 8][0] == board[cell_index][0]
                    for i in range(1, 4)
                ):
                    if board[cell_index][0] == "y":
                        return "yellow"
                    else:
                        return "red"

    # Überprüfung der Diagonalen (links unten nach rechts oben) Gewinnbedingung
    for row in range(3):
        for column in range(4):
            cell_index = column + 3 + row * 7
            if board[cell_index][0] != "n":
                if all(
                    board[cell_index + i * 6][0] == board[cell_index][0]
                    for i in range(1, 4)
                ):
                    if board[cell_index][0] == "y":
                        return "yellow"
                    else:
                        return "red"

    return won


def makeMove(board: list, move: int, color: str):
    """
    Führt einen Spielzug durch, indem ein Stein des Spielers auf dem Spielbrett platziert wird.

    Args:
        board: Eine Liste, die das Spielbrett repräsentiert.
        move: Der Index, in die der Spieler sein Stein setzen möchte (0-basiert).
        color: Die Farbe des Spielers (z.B. 'y' für gelb oder 'r' für rot).

    Returns:
        newBoard: Eine Kopie des ursprünglichen Spielbretts mit dem hinzugefügten Stein des Spielers.
    """
    # Kopiere das Spielbrett, um es nicht zu verändern
    newBoard = list(board)

    # Setze das Symbol des Spielers in der gewählten Spalte
    newBoard[move] = f"{color}{newBoard[move][1:]}"

    return newBoard


def moveSyntax(move: int):
    """
    Wandelt in ein gültiges Stringformat um

    Args:
        move: Repräsentiert den Zug

    Returns:
        Umgewandeltes Stringformat
    """
    if move < 10:
        move = "0" + str(move)
        return move
    else:
        return str(move)


# Die Spielfläche wird gezeichnet
def drawGame():
    for i in range(8):
        p.draw.line(
            screen,
            [255, 255, 255],
            [0, i * sizeOne + sizeOne - 2.5],
            [size, i * sizeOne + sizeOne - 2.5],
            5,
        )
        p.draw.line(
            screen, [255, 255, 255], [i * sizeOne, sizeOne], [i * sizeOne, size], 5
        )


# Ein gelder oder roter Kreis wird gezeichnet
def drawBall(x, y):
    if player:
        color = [255, 255, 0]
    else:
        color = [255, 0, 0]
    p.draw.circle(screen, color, [x, y], sizeOne / 2 - 5)


# Hier werden die Koordinaten geprüft und verarbeitet.
# Es wird nach dem nächst freien Feld ein einer Spalte gesucht.
# Schlussendlich werden die verarbeiteten Daten in den Listen occupied und board gespeichert.
def put(column: int, board: list):
    i = 35
    inColumn = False
    while i >= 0:
        if (i + column) not in occupied:
            occupied.append(i + column)
            p.draw.rect(screen, [0, 0, 0], (width / 2 - 75, height // 20 - 20, 150, 40))
            font = p.font.Font("freesansbold.ttf", 32)
            wait = font.render("Warten!", True, (0, 255, 0), (0, 0, 255))
            waitRect = wait.get_rect()
            waitRect.center = (width / 2, height // 20)
            screen.blit(wait, waitRect)
            p.display.update()
            time.sleep(0.5)
            if player:
                board[i + column] = "y" + moveSyntax(move)
            else:
                board[i + column] = "r" + moveSyntax(move)
            inColumn = True
            x = column * sizeOne + (sizeOne / 2)
            y = (i // 7) * sizeOne + (sizeOne * 1.5)
            drawBall(x, y)
            break
        else:
            i = i - 7
    return inColumn


# Boolean: True = (Das Spiel ist am laufen) False = (Das Spiel ist nicht mehr am laufen)
online = True

# Boolean: True = (Gelber Spieler) False = (Roter Spieler)
player = True

# Metadaten für den ersten Spieler (yellow)
yellowTime = None
yellowTotalTime = 0

# Metadaten für den zweiten Spieler (red)
redTime = None
redTotalTime = 0


# Solange das Spiel online ist, sollte es dauernd das Feld überprüfen.
while online:
    for event in p.event.get():
        if event.type == p.QUIT:
            # Spiel wird beendet
            online = False

        # Das Fenster wird neugeladen
        p.display.update()
        if event.type == p.MOUSEBUTTONDOWN:
            if player is not None:
                # In played wird gespeichert ob jemand gesetzt hat.
                column = p.mouse.get_pos()[0] // sizeOne
                played = put(column, board)
                # Spielzug wird um 1 erhöht
                move = move + 1
                # In status wird gespeichert ob jemand gewonnen hat.
                status = checkWin(board)
                if len(occupied) == 42:
                    draw = True
                else:
                    draw = False
                

            if played and player is not None:
                if not status and not draw:
                    if player:
                        # Spieler wird gewechselt
                        if yellowTime is not None:
                            yellowTotalTime = yellowTotalTime + (
                                time.time() - yellowTime
                            )
                        redTime = time.time()
                        p.draw.rect(
                            screen,
                            [0, 0, 0],
                            (width / 2 - 75, height // 20 - 20, 150, 40),
                        )
                        player = False
                    else:
                        # Spieler wird gewechselt
                        redTotalTime = redTotalTime + (time.time() - redTime)
                        yellowTime = time.time()
                        p.draw.rect(
                            screen,
                            [0, 0, 0],
                            (width / 2 - 75, height // 20 - 20, 150, 40),
                        )
                        player = True

                # Jemand hat gewonnen
                else:
                    p.draw.rect(
                        screen, [0, 0, 0], (width / 2 - 75, height // 20 - 20, 150, 40)
                    )
                    # Zeit des letzten Zuges wird addiert
                    if player:
                        yellowTotalTime = yellowTotalTime + (time.time() - yellowTime)
                    else:
                        redTotalTime = redTotalTime + (time.time() - redTime)

                    # Anzahl Züge werden berechnet
                    if move % 2 == 0:
                        yellowMoves = int(move / 2)
                    else:
                        yellowMoves = int(move / 2) + 1
                    redMoves = int(move / 2)

                    # Mappe abrufen / erstellen
                    try:
                        worksheet = workbook[sheet_name]

                    except KeyError:
                        workbook.create_sheet(sheet_name)
                        workbook.save(file_path_xlsx)
                        workbook = xl.load_workbook(file_path_xlsx)
                        worksheet = workbook[sheet_name]
                        setValue(1, 9, worksheet, "GameID:")
                        setStyle(1, 9, worksheet)
                        setValue(1, 10, worksheet, 0)
                        setStyle(1, 10, worksheet)
                        workbook.save(file_path_xlsx)
                        workbook = xl.load_workbook(file_path_xlsx)
                        worksheet = workbook[sheet_name]

                    # Aktuelle GameID abrufen
                    gameID = int(getValue(1, 10, worksheet))

                    # Alle Werte werden in Excel gespeichert
                    for i in range(6):
                        for k in range(7):
                            setValue(
                                14 * gameID + i + 1, k + 1, worksheet, board[i * 7 + k]
                            )
                            setStyle(
                                14 * gameID + i + 1, k + 1, worksheet
                            )  # (optional)

                    setValue(gameID * 14 + 9, 1, worksheet, "Yellow")
                    setValue(gameID * 14 + 10, 1, worksheet, "Red")

                    # Name
                    setValue(gameID * 14 + 8, 2, worksheet, "Name")
                    setValue(gameID * 14 + 9, 2, worksheet, yellowPlayer)
                    setValue(gameID * 14 + 10, 2, worksheet, redPlayer)

                    # Avarage Time
                    setValue(gameID * 14 + 8, 3, worksheet, "Avg Time")
                    setValue(
                        gameID * 14 + 9,
                        3,
                        worksheet,
                        round(yellowTotalTime / yellowMoves, 2),
                    )
                    setValue(
                        gameID * 14 + 10,
                        3,
                        worksheet,
                        round(redTotalTime / redMoves, 2),
                    )

                    # Accuracy
                    setValue(gameID * 14 + 8, 4, worksheet, "Accuracy")
                    setValue(gameID * 14 + 9, 4, worksheet, "None")
                    setValue(gameID * 14 + 10, 4, worksheet, "None")

                    # Blunder
                    setValue(gameID * 14 + 8, 5, worksheet, "Blunder")
                    setValue(gameID * 14 + 9, 5, worksheet, "None")
                    setValue(gameID * 14 + 10, 5, worksheet, "None")

                    # Date
                    setValue(gameID * 14 + 12, 1, worksheet, "Date")
                    setValue(gameID * 14 + 12, 2, worksheet, str(date.today()))

                    # Moves
                    setValue(gameID * 14 + 12, 4, worksheet, "Moves")
                    setValue(gameID * 14 + 12, 5, worksheet, move)

                    setValue(1, 10, worksheet, gameID + 1)
                    workbook.save(file_path_xlsx)

                    # Erster Parameter bestimmt die Schriftart
                    # Zweiter Parameter bestimmt die Schriftgrösse
                    font = p.font.Font("freesansbold.ttf", 32)

                    if not draw:
                        if player:
                            # Text und koordinaten werden bestimmt
                            text = font.render(
                                "Gelb hat gewonnen!", True, (0, 255, 0), (0, 0, 255)
                            )
                            textRect = text.get_rect()
                            textRect.center = (180, height // 20)
                        else:
                            # Text und koordinaten werden bestimmt
                            text = font.render(
                                "Rot hat gewonnen!", True, (0, 255, 0), (0, 0, 255)
                            )
                            textRect = text.get_rect()
                            textRect.center = (160, height // 20)
                    else:
                        # Text und koordinaten werden bestimmt
                            text = font.render(
                                "Unentschieden!", True, (0, 255, 0), (0, 0, 255)
                            )
                            textRect = text.get_rect()
                            textRect.center = (160, height // 20)

                    # player wird auf None gesetzt, damit man nicht mehr setzen kann.
                    player = None
                    # Text wird gesetzt
                    screen.blit(text, textRect)

                    text = font.render("Neustarten", True, (0, 255, 0), (0, 0, 255))
                    textRect = text.get_rect()
                    textRect.center = (width - 110, height // 20)
                    screen.blit(text, textRect)

            if player is None:
                position = p.mouse.get_pos()
                # Es wird überprüft wo ob man auf "Neustarten" gedürckt hat
                if textRect.x <= position[0] <= textRect.x + textRect.size[0]:
                    if textRect.y <= position[1] <= textRect.y + textRect.size[1]:
                        # Alle Werte werden zurückgesetzt
                        occupied = []
                        board = []
                        for n in range(42):
                            board.append("nnn")
                        move = 0
                        screen.fill(p.Color(0, 0, 0))
                        yellowTime = None
                        yellowTotalTime = 0
                        redTime = None
                        redTotalTime = 0
                        player = True

        # Das Feld wird gezeichnet
        drawGame()

p.quit()
