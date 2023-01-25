# Vier-Gewinnt Python-Programm (pygame) von Jery:

# Wichtig:
# Dieses Programm funktioniert nur auf MacOS!
# Es lässt sich auch ganz normal über Terminal starten.
# Ist der Code, welcher in der MacOS App steckt.

# Packages:
# -> https://github.com/JeryDev/MA_2022-2023/blob/main/source/MacOS/dependencies.txt

import pygame as p
import os
import sys
from os.path import dirname, join
import json
import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.styles.borders import BORDER_THIN
import platform

# Initialisierungsfunktion
p.init()

# Das Logo wird geladen
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = dirname(__file__)

    return join(base_path, relative_path)

try:
    icon = p.image.load(resource_path("images/icon.png"))
    p.display.set_icon(icon)

except(FileNotFoundError):
    pass


# Abfrage ob das Betriebssystem MacOS ist
if platform.system() == "Darwin":
    # Pfade der Dateien
    folder_path = os.path.join(os.environ["HOME"], "Desktop", "Maturaarbeit_Jery")
    file_path_xlsx = os.path.join(os.environ["HOME"], "Desktop", "Maturaarbeit_Jery", "data.xlsx")
    file_path_config = os.path.join(os.environ["HOME"], "Desktop", "Maturaarbeit_Jery", "config.json")
else:
    # Abbruch bei ungültigem Betriebssystem
    exit("Unsupported System")

# Werte aus Excel abrufen
def getValue(row, column, tabel):
    return tabel.cell(row=row, column=column).value

# Werte in Excel festlegen
def setValue(row: int, column: int, worksheet, value):
    worksheet.cell(row=row, column=column).value = value

# Überschaubares Design für einzelne Spalten setzen
def setStyle(row, column, worksheet):
    worksheet.cell(row=row, column=column).fill = PatternFill(fgColor="dadada", fill_type="solid")
    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='bfbfbf'),
        right=Side(border_style=BORDER_THIN, color='bfbfbf'),
        top=Side(border_style=BORDER_THIN, color='bfbfbf'),
        bottom=Side(border_style=BORDER_THIN, color='bfbfbf')
    )
    worksheet.cell(row=row, column=column).border = thin_border

# Überprüfen ob es den Ordner bereits gibt
# Wenn nicht dann einen neuen Ordner anlegen
if not os.path.isdir(folder_path):
    os.mkdir(folder_path)
                        
# Überprüfen ob es die Excel Datei bereits gibt
# Wenn nicht dann einen neuen Datei anlegen mit Standardwert
if not os.path.isfile(file_path_xlsx):
    workbook = xl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Stats"
    setValue(1, 9, worksheet, "GameID:")
    setStyle(1, 9, worksheet)
    setValue(1, 10, worksheet, 0)
    setStyle(1, 10, worksheet)
    workbook.save(file_path_xlsx)

# Überprüfen ob es die Konfigurationsdatei bereits gibt
# Wenn nicht dann einen neuen Datei anlegen mit Standardwert
if not os.path.isfile(file_path_config): 
    data = {
      "size": 1050
    }
    with open(file_path_config, "w") as file:
        json.dump(data, file, indent=1)
                                      

# Die Konfigurationsdatei öffnen
f = open(file_path_config)
  
# Konfigurationsdatei in ein Dictionary umwandeln
data = json.load(f)

# Den "Size" Parameter abrufen
size = data['size']
f.close()

# Grösse von einem Feld bestimmen
sizeOne = int(size / 7)

width = size
height = size

# Hier werden die Grössen eingetragen.
screen = p.display.set_mode((width, height))

# Fensterbeschriftung
p.display.set_caption("4-Gewinnt")

# Liste, welche die besetzten Felder speichert
occupied = []

# Liste, welche das aktuelle Feld speichert
field = []

# Leeres Feld wird generiert
for n in range(42):
    field.append("nnn")

#Aktueller Spielzug
move = 0

#Int wird in einen gültigen String umgewandelt
def moveSyntax(move):
    if move < 10:
        move = "0" + str(move)
        return move
    else:
        return str(move)

# Die Spielfläche wird gezeichnet
def drawGame():
    for i in range(8):
        p.draw.line(screen, [255, 255, 255], [0, i * sizeOne + sizeOne - 2.5], [size, i * sizeOne + sizeOne - 2.5], 5)
        p.draw.line(screen, [255, 255, 255], [i * sizeOne, sizeOne], [i * sizeOne, size], 5)


# Ein gelder oder roter Kreis wird gezeichnet
def drawBall(x, y):
    if player:
        color = [255, 255, 0]
    else:
        color = [255, 0, 0]
    p.draw.circle(screen, color, [x, y], sizeOne / 2 - 5)


# Hier werden die Koordinaten geprüft und verarbeitet.
# Es wird nach dem nächst freien Feld ein einer Spalte gesucht.
# Schlussendlich werden die verarbeiteten Daten in den Listen occupied und field gespeichert.
def put():
    column = p.mouse.get_pos()[0] // sizeOne
    i = 35
    inColumn = False
    while i >= 0:
        if (i + column) not in occupied:
            occupied.append(i + column)
            if player:
                field[i + column] = "y" + moveSyntax(move)
            else:
                field[i + column] = "r" + moveSyntax(move)
            inColumn = True
            x = column * sizeOne + (sizeOne / 2)
            y = (i // 7) * sizeOne + (sizeOne * 1.5)
            drawBall(x, y)
            break
        else:
            i = i - 7

    return inColumn


# Hier wird überprüft ob es bereits 4 gleiche Kreise in einer Reihe hat.
def checkWin():

    # Boolean: True = (Jemand hat gewonnen) False = (Niemand hat gewonnen)
    won = False

    # Überprüfung: Horizontal (rechts, links)
    for row in range(6):
        yellowInRow = 0
        redInRow = 0
        for column in range(7):
            if field[column + row * 7][0] == "r":
                redInRow = redInRow + 1
                yellowInRow = 0
            elif field[column + row * 7][0] == "y":
                yellowInRow = yellowInRow + 1
                redInRow = 0
            else:
                yellowInRow = 0
                redInRow = 0
            if redInRow == 4:
                won = True
                break
            if yellowInRow == 4:
                won = True
                break

    # Überprüfung: Vertikal (oben / unten)
    for column in range(7):
        yellowInRow = 0
        redInRow = 0
        for row in range(6):
            if field[column + row * 7][0] == "r":
                redInRow = redInRow + 1
                yellowInRow = 0
            elif field[column + row * 7][0] == "y":
                yellowInRow = yellowInRow + 1
                redInRow = 0
            else:
                yellowInRow = 0
                redInRow = 0
            if redInRow == 4:
                won = True
                break
            if yellowInRow == 4:
                won = True
                break

    # Überprüfung: Diagonal (rechts unten / links oben)
    for row in range(3):
        for column in range(4):
            yellowInRow = 0
            redInRow = 0
            for diagonal in range(4):
                if field[column + row * 7 + diagonal * 8][0] == "r":
                    redInRow = redInRow + 1
                    yellowInRow = 0
                elif field[column + row * 7 + diagonal * 8][0] == "y":
                    yellowInRow = yellowInRow + 1
                    redInRow = 0
                else:
                    yellowInRow = 0
                    redInRow = 0
                if redInRow == 4:
                    won = True
                    break
                if yellowInRow == 4:
                    won = True
                    break

    # Überprüfung: Diagonal (rechts oben / links unten)
    for row in range(3):
        for column in range(4):
            yellowInRow = 0
            redInRow = 0
            for diagonal in range(4):
                if field[column + 3 + row * 7 + diagonal * 6][0] == "r":
                    redInRow = redInRow + 1
                    yellowInRow = 0
                elif field[column + 3 + row * 7 + diagonal * 6][0] == "y":
                    yellowInRow = yellowInRow + 1
                    redInRow = 0
                else:
                    yellowInRow = 0
                    redInRow = 0
                if redInRow == 4:
                    won = True
                    break
                if yellowInRow == 4:
                    won = True
                    break
    return won


# Boolean: True = (Das Spiel ist am laufen) False = (Das Spiel ist nicht mehr am laufen)
online = True

# Boolean: True = (Gelber Spieler) False = (Roter Spieler)
player = True

# Solange das Spiel online ist, sollte es dauernd das Feld überprüfen.
while online:
    for event in p.event.get():
        if event.type == p.QUIT:
            # Spiel wird beendet
            online = False

        # Das Fenster wird neugeladen
        p.display.update()
        if event.type == p.MOUSEBUTTONDOWN:
            if player is not None:
                # In played wird gespeichert ob jemand gesetzt hat.
                played = put()
                #Spielzug wird um 1 erhöht
                move = move + 1
                # In status wird gespeichert ob jemand gewonnen hat.
                status = checkWin()

            if played and player is not None:
                if not status:
                    if player:
                        # Spieler wird gewechselt
                        player = False
                    else:
                        # Spieler wird gewechselt
                        player = True

                # Jemand hat gewonnen
                else:
                    # Excel Datei abrufen
                    workbook = xl.load_workbook(file_path_xlsx)
                    worksheet = workbook['Stats']

                    # Aktuelle GameID abrufen
                    gameID = getValue(1, 10, worksheet)

                    # Alle Werte in Excel speichern
                    for i in range(6):
                        for k in range(7):
                            setValue(12 * gameID + i + 1, k + 1, worksheet, field[i * 7 + k])
                            setStyle(12 * gameID + i + 1, k + 1, worksheet) #(optional)

                    setValue(gameID * 12 + 8, 1, worksheet, "Steps:")
                    setValue(gameID * 12 + 8, 2, worksheet, move)
                    setValue(gameID * 12 + 9, 1, worksheet, "R Value:")
                    setValue(gameID * 12 + 10, 1, worksheet, "Y Value:")
                    setValue(1, 10, worksheet, gameID + 1)
                    workbook.save(file_path_xlsx)
                    
                    # Erster Parameter bestimmt die Schriftart
                    # Zweiter Parameter bestimmt die Schriftgrösse
                    font = p.font.Font('freesansbold.ttf', 32)

                    if player:
                        # Text und koordinaten werden bestimmt
                        text = font.render('Gelb hat gewonnen!', True, (0, 255, 0), (0, 0, 255))
                        textRect = text.get_rect()
                        textRect.center = (180, height // 20)
                    else:
                        # Text und koordinaten werden bestimmt
                        text = font.render('Rot hat gewonnen!', True, (0, 255, 0), (0, 0, 255))
                        textRect = text.get_rect()
                        textRect.center = (160, height // 20)

                    # player wird auf None gesetzt, damit man nicht mehr setzen kann.
                    player = None
                    # Text wird gesetzt
                    screen.blit(text, textRect)

                    text = font.render('Neustarten', True, (0, 255, 0), (0, 0, 255))
                    textRect = text.get_rect()
                    textRect.center = (width - 110, height // 20)
                    screen.blit(text, textRect)

            if player is None:
                position = p.mouse.get_pos()
                # Es wird überprüft wo ob man auf "Neustarten" gedürckt hat
                if textRect.x <= position[0] <= textRect.x + textRect.size[0]:
                    if textRect.y <= position[1] <= textRect.y + textRect.size[1]:
                        # Alle Werte werden zurückgesetzt, damit das Spiel von vorne beginnt
                        occupied = []
                        field = []
                        for n in range(42):
                            field.append("nnn")
                        move = 0
                        screen.fill(p.Color(0, 0, 0))
                        player = True

        # Das Feld wird gezeichnet
        drawGame()

p.quit()
