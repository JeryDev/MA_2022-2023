import openpyxl as xl

path = "Pfad der Datendatei"
sheet_name = "Spieler1 vs Spieler2"
gameID = 0

# Werte aus Excel abrufen
def getValue(row: int, column: int, tabel):
    return tabel.cell(row=row, column=column).value

# Werte in Excel festlegen
def setValue(row: int, column: int, worksheet, value):
    worksheet.cell(row=row, column=column).value = value

def boardSyntax(index):
    if int(index) < 10:
        return f"0{index}"
    else:
        return str(index)


def difference(num1: int, num2: int):
    """
    Berechnet die Differenz von zwei Zahlen

    Args:
        num1: Erste Zahl
        num2: Zweite Zahl

    Returns:
        diff: Differenz der Zahlen
    """
    if num1 > num2:
        diff = num1 - num2
    else:
        diff = num2 - num1

    return diff


def getAccuracy(move: int, maxValue: int, minValue: int, color: str):
    """
    Überprüft die Qualität des Zuges

    Args:
        move: Repräsentiert den Wert des Zuges
        best: Repräsentiert den bestmöglichen Wert
        worst: Repräsentiert den schlechtesten Wert

    Returns:
        accuracy: Qualität des Zuges
    """
    if color == "y":
        spectrum = difference(maxValue, minValue)
        value = difference(move, minValue)
        if spectrum > 0:
            accuracy = 100 / spectrum * value
        else:
            accuracy = 100
        return round(accuracy, 1)
    elif color == "r":
        spectrum = difference(maxValue, minValue)
        value = difference(move, maxValue)
        if spectrum > 0:
            accuracy = 100 / spectrum * value
        else:
            accuracy = 100
        return round(accuracy, 1)


def isBlunder(accuracy, maxValue, minValue, move, color):
    """
    Überpüft, ob ein Zug ein Blunder ist

    Args:
        accuracy: Die Qualität des Zuges in Prozent
        maxValue: Der grösste Wert von allen möglichen Zügen
        minValue: Der kleinste Wert von allen möglichen Zügen
        move: Repräsentiert den Wert des Zuges (Nicht in Prozent)

    Returns:
        0: Kein Blunder
        1: Blunder
    """
    spectrum = difference(maxValue, minValue)

    # Verpasster Gewinn (Gelb)
    if color == "y" and maxValue >= 9996 and move < 9996:
        return 1
    # Verpasster Gewinn (Rot)
    if color == "r" and minValue <= -9996 and move > -9996:
        return 1
    # Gewinnchance für Rot
    if color == "y" and minValue <= -9996 and maxValue > -9996 and move <= -9996:
        return 1
    # Gewinnchance für Gelb
    if color == "r" and maxValue >= 9996 and minValue < 9996 and move >= 9996:
        return 1
    # Verpasster Zug
    if accuracy < 10 and spectrum >= 200:
        return 1
    # Kein Blunder
    else:
        return 0


def isValidDirection(next_index: int, end: int, direction: str):
    """
    Überprüft, ob ein Index in der gegebenen Richtung gültig ist, wenn er als
    nächstes besucht wird.

    Args:
        next_index: Der nächste Index, der überprüft werden soll.
        end: Der Index, von dem aus gestartet wird.
        direction: Die Richtung, in der überprüft werden soll.

    Returns:
        True, wenn der nächste Index in der gegebenen Richtung gültig ist, andernfalls False.
    """
    if direction == "vertical":
        if getRow(end) + 1 == getRow(next_index):
            # Der nächste Index ist eine Zeile darunter
            return True
        elif getRow(end) - 1 == getRow(next_index):
            # Der nächste Index ist eine Zeile darüber
            return True
    elif direction == "horizontal":
        if getColumn(end) + 1 == getColumn(next_index):
            # Der nächste Index ist eine Spalte rechts
            return True
        elif getColumn(end) - 1 == getColumn(next_index):
            # Der nächste Index ist eine Spalte links
            return True
    elif direction == "diagonal":
        if getRow(end) + 1 == getRow(next_index) and getColumn(end) + 1 == getColumn(
            next_index
        ):
            # Der nächste Index ist diagonal unten rechts
            return True
        elif getRow(end) - 1 == getRow(next_index) and getColumn(end) - 1 == getColumn(
            next_index
        ):
            # Der nächste Index ist diagonal oben links
            return True
        elif getRow(end) + 1 == getRow(next_index) and getColumn(end) - 1 == getColumn(
            next_index
        ):
            # Der nächste Index ist diagonal unten links
            return True
        elif getRow(end) - 1 == getRow(next_index) and getColumn(end) + 1 == getColumn(
            next_index
        ):
            # Der nächste Index ist diagonal oben rechts
            return True
    # Der nächste Index ist in keiner der erlaubten Richtungen gültig.
    return False


def getRow(index: int):
    """
    Gibt die Zeile zurück, in der sich der gegebene Index befindet.

    Args:
        index: Der Index, dessen Zeile zurückgegeben werden soll.

    Returns:
        Die Zeile, in der sich der gegebene Index befindet.
    """
    return index // 7 + 1


def getColumn(index: int):
    """
    Gibt die Spalte zurück, in der sich der gegebene Index befindet.

    Args:
        index: Der Index, dessen Spalte zurückgegeben werden soll.

    Returns:
        Die Spalte, in der sich der gegebene Index befindet.
    """
    return index % 7 + 1


def getColor(index: int, board: list):
    """
    Gibt die Farbe des Steins am gegebenen Index auf dem gegebenen Spielbrett zurück.

    Args:
        index: Der Index, an dem sich der Stein befindet.
        board: Eine Liste, die das Spielbrett repräsentiert.

    Returns:
        Die Farbe des Steins am gegebenen Index auf dem Spielbrett.
        Gibt "n" zurück, wenn kein Stein an diesem Index vorhanden ist.
    """
    if index < len(
        board
    ):  # Überprüfen, ob der gegebene Index im Bereich des Spielbretts liegt
        return board[index][0]
    else:
        return "n"


def getUnoccupiedLocations(board: list):
    """
    Gibt eine Liste von Indexe zurück, die auf freie Plätze auf dem Spielbrett verweisen.

    Args:
        board: Eine Liste, die das Spielbrett repräsentiert.

    Returns:
        Eine Liste von Indexe, die auf freie Plätze auf dem Spielbrett verweisen.
    """
    free_spaces = []
    for i in range(7):
        for k in range(6):
            # Berechnung des Index des Elements in der aktuellen Spalte und Zeile.
            v = i + ((5 - k) * 7)
            if board[v][0] == "n":
                free_spaces.append(v)
                break
    return free_spaces


def isMiddle(index: int):
    """
    Überprüft, ob der gegebene Index in der Mitte auf dem Spielbrett ist.

    Args:
        index: Der Index, der überprüft werden soll.

    Returns:
        True, wenn der gegebene Index in der Mitte auf dem Spielbrett ist.
        False, wenn der gegebene Index nicht in der Mitte auf dem Spielbrett ist.
    """
    return index in (3, 10, 17, 24, 31, 38)


DIRECTION_LOOKUP = {
    -8: (-1, -1, "diagonal"),
    -7: (-1, 0, "vertical"),
    -6: (-1, 1, "diagonal"),
    -1: (0, -1, "horizontal"),
    1: (0, 1, "horizontal"),
    6: (1, -1, "diagonal"),
    7: (1, 0, "vertical"),
    8: (1, 1, "diagonal"),
}


def getDirection(start: int, end: int):
    """
    Bestimmt die Richtung basierend auf dem Start- und Endindex.

    Args:
        start: Der Startindex.
        end: Der Endindex.

    Returns:
        Der nächste Index in der berechneten Richtung, wenn die Richtung gültig ist.
        False, wenn die Richtung ungültig ist oder der nächste Index außerhalb des Spielbretts liegt.
    """
    diff = end - start  # Die Differenz zwischen Start- und Endindex wird berechnet.
    # Wenn die Differenz in der DIRECTION_LOOKUP-Liste enthalten ist, gibt es eine gültige Richtung.
    if diff in DIRECTION_LOOKUP:
        # Die Zeilen- und Spaltenunterschiede sowie die Richtung werden aus der DIRECTION_LOOKUP-Liste abgerufen.
        row_diff, col_diff, direction = DIRECTION_LOOKUP[diff]
        # Der nächste Index in der berechneten Richtung wird berechnet.
        next_index = end + row_diff * 7 + col_diff
        # Überprüfung, ob der nächste Index in der berechneten Richtung gültig ist und innerhalb des Spielbretts liegt.
        if isValidDirection(next_index, end, direction) and 0 <= next_index <= 41:
            return next_index  # Der nächste Index wird zurückgegeben.
    # Wenn die Richtung ungültig ist oder der nächste Index außerhalb des Spielbretts liegt, wird False zurückgegeben.
    return False


def translateBoard(board):
    """
    Wandelt das Brett in ein neues Format um

    Args:
        board: Aktuelles Brett

    Returns:
        Umgewandltes Brett
    """
    rl = []
    for i in range(len(board)):
        rl.append(board[i][0])

    tl = []
    for i in range(6):
        v1 = 7 * i
        v2 = 7 * i + 7
        tl.append(rl[v1:v2])

    return tl


def getScore(window: str):
    countY = window.count("y")
    countR = window.count("r")
    countN = window.count("n")

    if countY + countN == 4 and countY > 0:
        if countY == 1:
            return 1
        elif countY == 2:
            return 10
        elif countY == 3:
            return 100
        else:
            return 10000
    if countR + countN == 4 and countR > 0:
        if countR == 1:
            return -1
        elif countR == 2:
            return -10
        elif countR == 3:
            return -100
        else:
            return -10000
    return 0

def evaluate(board: list):
    translatedBoard = translateBoard(board)

    # Wertung des Spielbretts
    score = 0

    # Horizontal
    for row in range(6):
        for col in range(4):
            window = translatedBoard[row][col : col + 4]
            window_string = "".join(window)
            score += getScore(window_string)

    # Vertikal
    for col in range(7):
        for row in range(3):
            window = [translatedBoard[row + i][col] for i in range(4)]
            window_string = "".join(window)
            score += getScore(window_string)

    # Diagonal (von links oben nach rechts unten)
    for row in range(3):
        for col in range(4):
            window = [translatedBoard[row + i][col + i] for i in range(4)]
            window_string = "".join(window)
            score += getScore(window_string)

    # Diagonal (von rechts oben nach links unten)
    for row in range(3):
        for col in range(3, 7):
            window = [translatedBoard[row + i][col - i] for i in range(4)]
            window_string = "".join(window)
            score += getScore(window_string)

    return score


def checkWin(board: list):
    """
    Überprüft ob jemand gewonnen hat

    Args:
        board: Eine Liste, die das Spielbrett repräsentiert.

    Returns:
        yellow: Gelb hat gewonnen.
        red: Rot hat gewonnen.
        None: Niemand hat gewonnen.
    """
    won = None

    # Überprüfung der horizontalen Gewinnbedingung
    for row in range(6):
        for column in range(4):
            cell_index = column + row * 7
            if board[cell_index][0] != "n":
                if all(
                    board[cell_index + i][0] == board[cell_index][0]
                    for i in range(1, 4)
                ):
                    if board[cell_index][0] == "y":
                        return "yellow"
                    else:
                        return "red"

    # Überprüfung der vertikalen Gewinnbedingung
    for column in range(7):
        for row in range(3):
            cell_index = column + row * 7
            if board[cell_index][0] != "n":
                if all(
                    board[cell_index + i * 7][0] == board[cell_index][0]
                    for i in range(1, 4)
                ):
                    if board[cell_index][0] == "y":
                        return "yellow"
                    else:
                        return "red"

    # Überprüfung der Diagonalen (rechts unten nach links oben) Gewinnbedingung
    for row in range(3):
        for column in range(4):
            cell_index = column + row * 7
            if board[cell_index][0] != "n":
                if all(
                    board[cell_index + i * 8][0] == board[cell_index][0]
                    for i in range(1, 4)
                ):
                    if board[cell_index][0] == "y":
                        return "yellow"
                    else:
                        return "red"

    # Überprüfung der Diagonalen (links unten nach rechts oben) Gewinnbedingung
    for row in range(3):
        for column in range(4):
            cell_index = column + 3 + row * 7
            if board[cell_index][0] != "n":
                if all(
                    board[cell_index + i * 6][0] == board[cell_index][0]
                    for i in range(1, 4)
                ):
                    if board[cell_index][0] == "y":
                        return "yellow"
                    else:
                        return "red"

    return won


def makeMove(board: list, move: int, color: str):
    """
    Führt einen Spielzug durch, indem ein Stein des Spielers auf dem Spielbrett platziert wird.

    Args:
        board: Eine Liste, die das Spielbrett repräsentiert.
        move: Der Index, in die der Spieler sein Stein setzen möchte (0-basiert).
        color: Die Farbe des Spielers (z.B. 'y' für gelb oder 'r' für rot).

    Returns:
        newBoard: Eine Kopie des ursprünglichen Spielbretts mit dem hinzugefügten Stein des Spielers.
    """
    # Kopiere das Spielbrett, um es nicht zu verändern
    newBoard = list(board)

    # Setze das Symbol des Spielers in der gewählten Spalte
    newBoard[move] = f"{color}{newBoard[move][1:]}"

    return newBoard


def getResults(board: list, depth: int, color: str, alpha: int, beta: int):
    """
    Wandelt das Spielbrett in ein Format um, das von der Minimax-Funktion verwendet werden kann.

    Args:
        board: Eine Liste, die das Spielbrett repräsentiert.
        depth: Die maximale Tiefe der Rekursion, die vom Minimax-Algorithmus durchlaufen werden soll.
        color: Die Farbe des Spielers, dessen bester Spielzug berechnet werden soll (z.B. 'y' für gelb oder 'r' für rot).
        alpha: Der aktuelle Alpha-Wert des Alpha-Beta-Pruning-Algorithmus.
        beta: Der aktuelle Beta-Wert des Alpha-Beta-Pruning-Algorithmus.

    Returns:
        Eine Liste mit dem besten Spielzug und seinem Bewertungswert, wenn die maximale Rekursionstiefe erreicht wurde.
        Andernfalls eine Liste mit den Bewertungswerten aller möglichen Spielzüge.
    """
    # Wenn die maximale Rekursionstiefe erreicht wurde, gebe False zurück
    if depth <= 0:
        return False

    # Initialisiere eine leere Ergebnisliste
    results = []

    # Durchlaufe alle freien Positionen auf dem Spielbrett
    for position in getUnoccupiedLocations(board):
        # Führe den Spielzug aus und berechne den Bewertungswert des Spielbretts
        newBoard = makeMove(board, position, color)
        if color == "y":
            maxPlayer = False
            value = minimax(newBoard, depth - 1, maxPlayer, alpha, beta)
            # Füge einen Bonus hinzu, wenn der Spielzug in der Mitte der untersten Reihe erfolgt
            if isMiddle(position):
                value = value + 4
        elif color == "r":
            maxPlayer = True
            value = minimax(newBoard, depth - 1, maxPlayer, alpha, beta)
            # Ziehe einen Malus ab, wenn der Spielzug in der Mitte der untersten Reihe erfolgt
            if isMiddle(position):
                value = value - 4

        # Füge den Bewertungswert des Spielzugs und seine Position der Ergebnisliste hinzu
        results.append([value, position])

    # Gibt die Ergebnisliste zurück
    values = []
    for i in range(len(results)):
        values.append(results[i][0])
    positions = []
    for i in range(len(results)):
        positions.append(results[i][1])
    return values, positions


def minimax(board: list, depth: int, maxPlayer: bool, alpha: int, beta: int):
    """
    Ein Algorithmus, der den bestmöglichen Spielzug in einer Connect-4-Partie findet.

    Args:
        board: Eine Liste, die das Spielbrett repräsentiert.
        position: Die Position des letzten Zugs.
        depth: Die aktuelle Tiefe im Suchbaum.
        maxPlayer: Ein Boolean, der angibt, ob der maximierende Spieler am Zug ist.
        alpha: Der aktuelle Alpha-Wert.
        beta: Der aktuelle Beta-Wert.

    Returns:
        Der Wert des besten Spielzugs.
    """

    # Wenn die maximale Tiefe erreicht wurde oder das Spiel gewonnen wurde
    winner = checkWin(board)
    if depth == 0 or winner != None:
        # Falls Gelb gewonnen hat, gib 10000 zurück
        if winner == "yellow":
            return 10000
        # Falls Rot gewonnen hat, gib -10000 zurück
        elif winner == "red":
            return -10000
        # Falls die maximale Tiefe erreicht wurde, berechne den Wert des aktuellen Boards und gib ihn zurück
        return evaluate(board)

    # Wenn der maximierende Spieler am Zug ist
    if maxPlayer:
        bestValue = -1000000
        # Für jeden möglichen Zug
        for move in getUnoccupiedLocations(board):
            # Mach den Zug auf dem Board
            newBoard = makeMove(board, move, "y")
            # Berechne den Wert des Zuges mit Hilfe des Minimax-Algorithmus
            value = minimax(newBoard, depth - 1, False, alpha, beta)
            # Wenn der berechnete Wert besser als der aktuelle beste Wert ist, aktualisiere den besten Wert
            bestValue = max(bestValue, value)
            # Aktualisiere den Alpha-Wert
            alpha = max(alpha, bestValue)
            # Wenn der Beta-Wert kleiner als oder gleich dem Alpha-Wert ist, breche die Schleife ab
            if beta <= alpha:
                break
        # Gib den besten Wert zurück
        return bestValue

    # Wenn der minimierende Spieler am Zug ist
    else:
        bestValue = 1000000
        # Für jeden möglichen Zug
        for move in getUnoccupiedLocations(board):
            # Mach den Zug auf dem Board
            newBoard = makeMove(board, move, "r")
            # Berechne den Wert des Zuges mit Hilfe des Minimax-Algorithmus
            value = minimax(newBoard, depth - 1, True, alpha, beta)
            # Wenn der berechnete Wert besser als der aktuelle beste Wert ist, aktualisiere den besten Wert
            bestValue = min(bestValue, value)
            # Aktualisiere den Beta-Wert
            beta = min(beta, bestValue)
            # Wenn der Beta-Wert kleiner als oder gleich dem Alpha-Wert ist, breche die Schleife ab
            if beta <= alpha:
                break
        # Gib den besten Wert zurück
        return bestValue


# Int wird in einen gültigen String umgewandelt
def moveSyntax(move: int):
    """
    Wandelt in ein gültiges Stringformat um

    Args:
        move: Repräsentiert den Zug

    Returns:
        Umgewandeltes Stringformat
    """
    if move < 10:
        move = "0" + str(move)
        return move
    else:
        return str(move)


def update_progress(progress):
    bar_length = 50  # Länge der Fortschrittsanzeige
    filled_length = int(progress * bar_length)
    bar = "=" * filled_length + "-" * (bar_length - filled_length)
    percent = progress * 100
    print(f"[{bar}] {percent:.1f}% complete", end="\r")

workbook_data = xl.load_workbook(path)
worksheet_data = workbook_data[sheet_name]
moves = int(getValue(gameID * 14 + 12, 5, worksheet_data))

board_from_excel = []
for row in range(6):
    for column in range(7):
        board_from_excel.append(str(getValue((gameID * 14) + row + 1, column + 1, worksheet_data)))

translatedBoard = []
for i in range(len(board_from_excel)):
    translatedBoard.append(board_from_excel[i][1:3])

accuracy_yellow = 0
blunder_yellow = 0

accuracy_red = 0
blunder_red = 0

depth = 7

board = []
for n in range(42):
    board.append("nnn")

for i in range(moves):
    if i % 2 == 0:
        index = translatedBoard.index(boardSyntax(i))
        values, positions = getResults(board, depth, "y", -1000000, 1000000)
        current = values[positions.index(index)]
        accuracy = getAccuracy(current, max(values), min(values), "y")
        blunder = isBlunder(accuracy, max(values), min(values), current, "y")
        accuracy_yellow = accuracy_yellow + accuracy
        blunder_yellow = blunder_yellow + blunder
        board[index] = f"y{boardSyntax(i)}"
    else:
        index = translatedBoard.index(boardSyntax(i))
        values, positions = getResults(board, depth, "r", -1000000, 1000000)
        current = values[positions.index(index)]
        accuracy = getAccuracy(current, max(values), min(values), "r")
        blunder = isBlunder(accuracy, max(values), min(values), current, "r")
        accuracy_red = accuracy_red + accuracy
        blunder_red = blunder_red + blunder
        board[index] = f"r{boardSyntax(i)}"

    progress = (i + 1) / moves
    update_progress(progress)

accuracy_yellow = round(accuracy_yellow / (int(moves / 2) + 1), 2)
accuracy_red = round(accuracy_red / (int(moves / 2)), 2)

old_accuracy_yellow = getValue(gameID * 14 + 9, 4, worksheet_data)
old_blunder_yellow = getValue(gameID * 14 + 9, 5, worksheet_data)

old_accuracy_red = getValue(gameID * 14 + 10, 4, worksheet_data)
old_blunder_red = getValue(gameID * 14 + 10, 5, worksheet_data)

setValue(gameID * 14 + 9, 4, worksheet_data, accuracy_yellow)
setValue(gameID * 14 + 9, 5, worksheet_data, blunder_yellow)

setValue(gameID * 14 + 10, 4, worksheet_data, accuracy_red)
setValue(gameID * 14 + 10, 5, worksheet_data, blunder_red)
workbook_data.save(path)


print("\n")
print(f"Old Data:")
print(f"Accuracy Yellow: {old_accuracy_yellow}")
print(f"Blunder Yellow: {old_blunder_yellow}")
print(f"Accuracy Red: {old_accuracy_red}")
print(f"Blunder Red: {old_blunder_red}")
print("\n")
print(f"Converted Data:")
print(f"Accuracy Yellow: {accuracy_yellow}")
print(f"Blunder Yellow: {blunder_yellow}")
print(f"Accuracy Red: {accuracy_red}")
print(f"Blunder Red: {blunder_red}")
