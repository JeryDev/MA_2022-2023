#Abspeicherungsprogramm für Excel von Jery:
#In diesem Programm wird veranschaulicht wie eine Partie in Excel abgespeichert wird.
#Für dies wird die Implementation "openpyxl" genutzt.
#Link zur Dokumentation: https://openpyxl.readthedocs.io/en/stable/#

import os
import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.styles.borders import BORDER_THIN
import platform

# Dateityp: xlsx
if platform.system() == "Darwin":
    path = os.path.join(os.environ["HOME"], "Desktop", "Matura_Jery", "data.xlsx")
elif platform.system == "Windows":
    path = os.path.join(os.environ["HOMEPATH"], "Desktop", "Matura_Jery", "data.xlsx")
else:
    exit("Unsupported Sysetm")

workbook = xl.load_workbook(path)
worksheet = workbook['Stats']


def setValue(row, column, tabel, value):
    tabel.cell(row=row, column=column).value = value


#Die GetValue Methode braucht man später für die Daten Auswertung.
def getValue(row, column, tabel):
    return tabel.cell(row=row, column=column).value


#Diese Methode ist optional. Sie dient nur für ein überschaubares Design.
def setStyle(row, column, tabel):
    tabel.cell(row=row, column=column).fill = PatternFill(fgColor="dadada", fill_type="solid")
    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='bfbfbf'),
        right=Side(border_style=BORDER_THIN, color='bfbfbf'),
        top=Side(border_style=BORDER_THIN, color='bfbfbf'),
        bottom=Side(border_style=BORDER_THIN, color='bfbfbf')
    )
    tabel.cell(row=row, column=column).border = thin_border


#Die gameID muss bei jedem Eintrag verändert werden.
#Der kleinste Wert ist 0 und man kann so viele Partien eintragen wie man will.
#Bei jedem zusätzlichen Eintrag muss man den Wert um 1 erhöhen.
#Es werden 12 Zeilen und 7 Spalten pro GameID bearbeitet.
gameID = getValue(1, 10, worksheet)

#Liste welche das Feld abspeichert.
field = []

#Da wird die Anzahl der Züge gespeichert.
steps = 0

#Startliste wird generiert.
for i in range(42):
    field.append("nnn")

#Eintragen der Daten in Excel.
for i in range(6):
    for k in range(7):
        setValue(12 * gameID + i + 1, k + 1, worksheet, field[i * 7 + k])
        setStyle(12 * gameID + i + 1, k + 1, worksheet) #(optional)

setValue(gameID * 12 + 8, 1, worksheet, "Steps:")
setValue(gameID * 12 + 8, 2, worksheet, steps)
setValue(gameID * 12 + 9, 1, worksheet, "R Value:")
setValue(gameID * 12 + 10, 1, worksheet, "Y Value:")

#Excel Datei wird abgespeichert.
workbook.save(path)
