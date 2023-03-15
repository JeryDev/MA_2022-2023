# Vier-Gewinnt Python-Programm (pygame) von Jery:

# Wichtig:
# Dieses Programm funktioniert nur auf MacOS!
# Es lässt sich auch ganz normal über Terminal starten.

# Packages:
# -> https://github.com/JeryDev/MA_2022-2023/blob/main/source/MacOS/dependencies.txt

import pygame as p
import os
import json
import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.styles.borders import BORDER_THIN
import platform

# Initialisierungsfunktion
p.init()


def is_valid_direction(next_index: int, end: int, direction: str):
    if direction == "vertical":
        if get_row(end) + 1 == get_row(next_index):
            return True
        elif get_row(end) - 1 == get_row(next_index):
            return True
    elif direction == "horizontal":
        if get_column(end) + 1 == get_column(next_index):
            return True
        elif get_column(end) - 1 == get_column(next_index):
            return True
    elif direction == "diagonal":
        if get_row(end) + 1 == get_row(next_index) and get_column(end) + 1 == get_column(next_index):
            return True
        elif get_row(end) - 1 == get_row(next_index) and get_column(end) - 1 == get_column(next_index):
            return True
        elif get_row(end) + 1 == get_row(next_index) and get_column(end) - 1 == get_column(next_index):
            return True
        elif get_row(end) - 1 == get_row(next_index) and get_column(end) + 1 == get_column(next_index):
            return True
    return False


def get_row(index: int):
    return index // 7 + 1


def get_column(index: int):
    return index % 7 + 1


def get_Color(index: int, board: list):
    if index < len(board):
        return board[index][0]
    else:
        return "n"


def get_free_spaces(board: list):
    free_spaces = []
    for i in range(7):
        for k in range(6):
            v = i + ((5 - k) * 7)
            if board[v][0] == "n":
                free_spaces.append(v)
                break
    return free_spaces


def check_middle_position(index: int):
    return index in (3, 10, 17, 24, 31, 38)


DIRECTION_LOOKUP = {
    -8: (-1, -1, "diagonal"), -7: (-1, 0, "vertical"), -6: (-1, 1, "diagonal"),
    -1: (0, -1, "horizontal"),                          1: (0, 1, "horizontal"),
    6: (1, -1, "diagonal"),  7: (1, 0, "vertical"),   8: (1, 1, "diagonal")
}


def get_direction(start: int, end: int):
    diff = end - start
    if diff in DIRECTION_LOOKUP:
        row_diff, col_diff, direction = DIRECTION_LOOKUP[diff]
        next_index = end + row_diff * 7 + col_diff
        if is_valid_direction(next_index, end, direction) and 0 <= next_index <= 41:
            return next_index
    return False


def get_all_neighbours(index: int):
    all_neighbours = []
    right = False
    left = False
    if index // 7 == (index + 1) // 7:
        right = True
    if index // 7 == (index - 1) // 7:
        left = True
    if (index - 7) // 7 >= 0:
        all_neighbours.append(index - 7)
        if right == True:
            all_neighbours.append(index - 6)
        if left == True:
            all_neighbours.append(index - 8)
    if (index + 7) // 7 <= 5:
        all_neighbours.append(index + 7)
        if right == True:
            all_neighbours.append(index + 8)
        if left == True:
            all_neighbours.append(index + 6)
    if right == True:
        all_neighbours.append(index + 1)
    if left == True:
        all_neighbours.append(index - 1)

    all_neighbours.sort()
    return all_neighbours


def analyse(board: list, position: int, color: str):
    value = 0
    all_neighbours = get_all_neighbours(position)
    for i in range(len(all_neighbours)):
        get_color_of_neighbour = get_Color(all_neighbours[i], board)
        if get_color_of_neighbour == color:
            value = value + 10
            neighbour = get_direction(position, all_neighbours[i])
            if neighbour != False:
                get_color_of_neighbour = get_Color(neighbour, board)
                if get_color_of_neighbour == color:
                    value = value + 100
                    neighbour = get_direction(all_neighbours[i], neighbour)
                    if neighbour != False:
                        get_color_of_neighbour = get_Color(neighbour, board)
                        if get_color_of_neighbour == color:
                            value = 10000
                        else:
                            neighbour = get_direction(
                                all_neighbours[i], position)
                            if neighbour != False:
                                get_color_of_neighbour = get_Color(
                                    neighbour, board)
                                if get_all_neighbours == color:
                                    value = 10000
                    else:
                        neighbour = get_direction(all_neighbours[i], position)
                        if neighbour != False:
                            get_color_of_neighbour = get_Color(
                                neighbour, board)
                            if get_color_of_neighbour == color:
                                value = 10000
                else:
                    neighbour = get_direction(all_neighbours[i], position)
                    if neighbour != False:
                        get_color_of_neighbour = get_Color(neighbour, board)
                        if get_color_of_neighbour == color:
                            value = value + 100
                            neighbour = get_direction(position, neighbour)
                            if neighbour != False:
                                get_color_of_neighbour = get_Color(
                                    neighbour, board)
                                if get_color_of_neighbour == color:
                                    value = 10000
            else:
                neighbour = get_direction(all_neighbours[i], position)
                if neighbour != False:
                    get_color_of_neighbour = get_Color(neighbour, board)
                    if get_color_of_neighbour == color:
                        value = value + 100
                        neighbour = get_direction(position, neighbour)
                        if neighbour != False:
                            get_color_of_neighbour = get_Color(
                                neighbour, board)
                            if get_color_of_neighbour == color:
                                value = 10000

    if color == "y":
        return value
    else:
        return -value


def check_win(board: list):
    won = None
    for row in range(6):
        yellowInRow = 0
        redInRow = 0
        for column in range(7):
            if board[column + row * 7][0] == "r":
                redInRow = redInRow + 1
                yellowInRow = 0
            elif board[column + row * 7][0] == "y":
                yellowInRow = yellowInRow + 1
                redInRow = 0
            else:
                yellowInRow = 0
                redInRow = 0
            if redInRow == 4:
                won = "red"
                break
            if yellowInRow == 4:
                won = "yellow"
                break

    # Überprüfung: Vertikal (oben / unten)
    for column in range(7):
        yellowInRow = 0
        redInRow = 0
        for row in range(6):
            if board[column + row * 7][0] == "r":
                redInRow = redInRow + 1
                yellowInRow = 0
            elif board[column + row * 7][0] == "y":
                yellowInRow = yellowInRow + 1
                redInRow = 0
            else:
                yellowInRow = 0
                redInRow = 0
            if redInRow == 4:
                won = "red"
                break
            if yellowInRow == 4:
                won = "yellow"
                break

    # Überprüfung: Diagonal (rechts unten / links oben)
    for row in range(3):
        for column in range(4):
            yellowInRow = 0
            redInRow = 0
            for diagonal in range(4):
                if board[column + row * 7 + diagonal * 8][0] == "r":
                    redInRow = redInRow + 1
                    yellowInRow = 0
                elif board[column + row * 7 + diagonal * 8][0] == "y":
                    yellowInRow = yellowInRow + 1
                    redInRow = 0
                else:
                    yellowInRow = 0
                    redInRow = 0
                if redInRow == 4:
                    won = "red"
                    break
                if yellowInRow == 4:
                    won = "yellow"
                    break

    # Überprüfung: Diagonal (rechts oben / links unten)
    for row in range(3):
        for column in range(4):
            yellowInRow = 0
            redInRow = 0
            for diagonal in range(4):
                if board[column + 3 + row * 7 + diagonal * 6][0] == "r":
                    redInRow = redInRow + 1
                    yellowInRow = 0
                elif board[column + 3 + row * 7 + diagonal * 6][0] == "y":
                    yellowInRow = yellowInRow + 1
                    redInRow = 0
                else:
                    yellowInRow = 0
                    redInRow = 0
                if redInRow == 4:
                    won = "red"
                    break
                if yellowInRow == 4:
                    won = "yellow"
                    break
    return won


def make_move(board: list, move: int, color: str):
    newBoard = []
    for i in range(len(board)):
        newBoard.append(board[i])
    newBoard[move] = f"{color}{newBoard[move][1:]}"
    return newBoard


def to_minimax(board: list, depth: int, color: str, alpha: int, beta: int):
    if depth > 0:
        results = []
        for position in get_free_spaces(board):
            newBoard = make_move(board, position, color)
            if color == "y":
                maxPlayer = False
                value = minimax(newBoard, position, depth -
                                1, maxPlayer, alpha, beta)
                if check_middle_position(position):
                    value = value + 4
            elif color == "r":
                maxPlayer = True
                value = minimax(newBoard, position, depth -
                                1, maxPlayer, alpha, beta)
                if check_middle_position(position):
                    value = value - 4
            results.append([value, position])
        print(results)
        if color == "y":
            max = [-1000000, None]
            for res in results:
                if res[0] >= max[0]:
                    max = res
            return max

        if color == "r":
            min = [1000000, None]
            for res in results:
                if res[0] <= min[0]:
                    min = res
            return min

        return results
    return False


def minimax(board: list, position: int, depth: int, maxPlayer: bool, alpha: int, beta: int):
    if depth == 0 or check_win(board) != None:
        if check_win(board) == "yellow":
            return 10000
        elif check_win(board) == "red":
            return -10000
        if depth == 0:
            if maxPlayer:
                return analyse(board, position, "r")
            else:
                return analyse(board, position, "y")

    if maxPlayer:
        bestValue = -1000000
        for move in get_free_spaces(board):
            newBoard = make_move(board, move, "y")
            value = minimax(newBoard, move, depth - 1, False, alpha, beta)
            bestValue = max(bestValue, value)
            alpha = max(alpha, bestValue)
            if beta <= alpha:
                break
        return bestValue

    else:
        bestValue = 1000000
        for move in get_free_spaces(board):
            newBoard = make_move(board, move, "r")
            value = minimax(newBoard, move, depth - 1, True, alpha, beta)
            bestValue = min(bestValue, value)
            beta = min(beta, bestValue)
            if beta <= alpha:
                break
        return bestValue


# Abfrage ob das Betriebssystem MacOS ist
if platform.system() == "Darwin":
    # Pfade der Dateien
    folder_path = os.path.join(
        os.environ["HOME"], "Desktop", "Maturaarbeit_Jery")
    file_path_xlsx = os.path.join(
        os.environ["HOME"], "Desktop", "Maturaarbeit_Jery", "data.xlsx")
    file_path_config = os.path.join(
        os.environ["HOME"], "Desktop", "Maturaarbeit_Jery", "config.json")
else:
    # Abbruch bei ungültigem Betriebssystem
    exit("Unsupported System")

# Werte aus Excel abrufen


def getValue(row, column, tabel):
    return tabel.cell(row=row, column=column).value

# Werte in Excel festlegen


def setValue(row: int, column: int, worksheet, value):
    worksheet.cell(row=row, column=column).value = value

# Überschaubares Design für einzelne Spalten setzen


def setStyle(row, column, worksheet):
    worksheet.cell(row=row, column=column).fill = PatternFill(
        fgColor="dadada", fill_type="solid")
    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='bfbfbf'),
        right=Side(border_style=BORDER_THIN, color='bfbfbf'),
        top=Side(border_style=BORDER_THIN, color='bfbfbf'),
        bottom=Side(border_style=BORDER_THIN, color='bfbfbf')
    )
    worksheet.cell(row=row, column=column).border = thin_border


# Überprüfen ob es den Ordner bereits gibt
# Wenn nicht dann einen neuen Ordner anlegen
if not os.path.isdir(folder_path):
    os.mkdir(folder_path)

# Überprüfen ob es die Excel Datei bereits gibt
# Wenn nicht dann einen neuen Datei anlegen mit Standardwert
if not os.path.isfile(file_path_xlsx):
    workbook = xl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Stats"
    setValue(1, 9, worksheet, "GameID:")
    setStyle(1, 9, worksheet)
    setValue(1, 10, worksheet, 0)
    setStyle(1, 10, worksheet)
    workbook.save(file_path_xlsx)

# Überprüfen ob es die Konfigurationsdatei bereits gibt
# Wenn nicht dann einen neuen Datei anlegen mit Standardwert
if not os.path.isfile(file_path_config):
    data = {
        "size": 1050
    }
    with open(file_path_config, "w") as file:
        json.dump(data, file, indent=1)


# Die Konfigurationsdatei öffnen
f = open(file_path_config)

# Konfigurationsdatei in ein Dictionary umwandeln
data = json.load(f)

# Den "Size" Parameter abrufen
size = data['size']
f.close()

# Grösse von einem Feld bestimmen
sizeOne = int(size / 7)

width = size
height = size

# Hier werden die Grössen eingetragen.
screen = p.display.set_mode((width, height))

# Fensterbeschriftung
p.display.set_caption("Vier-Gewinnt")

# Liste, welche die besetzten Felder speichert
occupied = []

# Liste, welche das aktuelle Feld speichert
board = []

# Leeres Feld wird generiert
for n in range(42):
    board.append("nnn")

# Aktueller Spielzug
move = 0

# Int wird in einen gültigen String umgewandelt


def moveSyntax(move):
    if move < 10:
        move = "0" + str(move)
        return move
    else:
        return str(move)

# Die Spielfläche wird gezeichnet


def drawGame():
    for i in range(8):
        p.draw.line(screen, [255, 255, 255], [
                    0, i * sizeOne + sizeOne - 2.5], [size, i * sizeOne + sizeOne - 2.5], 5)
        p.draw.line(screen, [255, 255, 255], [
                    i * sizeOne, sizeOne], [i * sizeOne, size], 5)


# Ein gelder oder roter Kreis wird gezeichnet
def drawBall(x, y):
    if player:
        color = [255, 255, 0]
    else:
        color = [255, 0, 0]
    p.draw.circle(screen, color, [x, y], sizeOne / 2 - 5)


# Hier werden die Koordinaten geprüft und verarbeitet.
# Es wird nach dem nächst freien Feld ein einer Spalte gesucht.
# Schlussendlich werden die verarbeiteten Daten in den Listen occupied und board gespeichert.
def put(column: int):
    i = 35
    inColumn = False
    while i >= 0:
        if (i + column) not in occupied:
            occupied.append(i + column)
            if player:
                board[i + column] = "y" + moveSyntax(move)
            else:
                board[i + column] = "r" + moveSyntax(move)
            inColumn = True
            x = column * sizeOne + (sizeOne / 2)
            y = (i // 7) * sizeOne + (sizeOne * 1.5)
            drawBall(x, y)
            break
        else:
            i = i - 7

    return inColumn


# Boolean: True = (Das Spiel ist am laufen) False = (Das Spiel ist nicht mehr am laufen)
online = True

# Boolean: True = (Gelber Spieler) False = (Roter Spieler)
player = True


def gameover(file_path_xlsx, board, player):
    # Excel Datei abrufen
    workbook = xl.load_workbook(file_path_xlsx)
    worksheet = workbook['Stats']

    # Aktuelle GameID abrufen
    gameID = getValue(1, 10, worksheet)

    # Alle Werte in Excel speichern
    for i in range(6):
        for k in range(7):
            setValue(12 * gameID + i + 1, k + 1, worksheet, board[i * 7 + k])
            setStyle(12 * gameID + i + 1, k + 1, worksheet)  # (optional)

    setValue(gameID * 12 + 8, 1, worksheet, "Steps:")
    setValue(gameID * 12 + 8, 2, worksheet, move)
    setValue(gameID * 12 + 9, 1, worksheet, "R Value:")
    setValue(gameID * 12 + 10, 1, worksheet, "Y Value:")
    setValue(1, 10, worksheet, gameID + 1)
    workbook.save(file_path_xlsx)

    # Erster Parameter bestimmt die Schriftart
    # Zweiter Parameter bestimmt die Schriftgrösse
    font = p.font.Font('freesansbold.ttf', 32)

    if player:
        # Text und koordinaten werden bestimmt
        text = font.render('Gelb hat gewonnen!', True,
                           (0, 255, 0), (0, 0, 255))
        textRect = text.get_rect()
        textRect.center = (180, height // 20)
    else:
        # Text und koordinaten werden bestimmt
        text = font.render('Rot hat gewonnen!', True, (0, 255, 0), (0, 0, 255))
        textRect = text.get_rect()
        textRect.center = (160, height // 20)

        # player wird auf None gesetzt, damit man nicht mehr setzen kann.
        player = None
        # Text wird gesetzt
        screen.blit(text, textRect)

        text = font.render('Neustarten', True, (0, 255, 0), (0, 0, 255))
        textRect = text.get_rect()
        textRect.center = (width - 110, height // 20)
        screen.blit(text, textRect)


# Solange das Spiel online ist, sollte es dauernd das Feld überprüfen.
while online:
    for event in p.event.get():
        if event.type == p.QUIT:
            # Spiel wird beendet
            online = False

        # Das Fenster wird neugeladen
        p.display.update()
        if len(occupied) == len(board):
            player = None
            font = p.font.Font('freesansbold.ttf', 32)
            text = font.render('Unentschieden!', True,
                               (0, 255, 0), (0, 0, 255))
            textRect = text.get_rect()
            textRect.center = (180, height // 20)
            screen.blit(text, textRect)
            text = font.render('Neustarten', True, (0, 255, 0), (0, 0, 255))
            textRect = text.get_rect()
            textRect.center = (width - 110, height // 20)
            screen.blit(text, textRect)

        if player == False:
            depth = 6
            result = to_minimax(board, depth, "r", -1000000, 1000000)
            occupied.append(result[1])
            board[result[1]] = f"r{moveSyntax(move)}"
            column = get_column(result[1]) - 1
            row = get_row(result[1]) - 1
            x = column * sizeOne + (sizeOne / 2)
            y = row * sizeOne + (sizeOne * 1.5)
            drawBall(x, y)
            move = move + 1
            status = check_win(board)
            if played and player is not None:
                if not status:
                    player = True
                    p.display.update()

                else:
                    p.display.update()
                    # Excel Datei abrufen
                    workbook = xl.load_workbook(file_path_xlsx)
                    worksheet = workbook['Stats']

                    # Aktuelle GameID abrufen
                    gameID = getValue(1, 10, worksheet)

                    # Alle Werte in Excel speichern
                    for i in range(6):
                        for k in range(7):
                            setValue(12 * gameID + i + 1, k + 1,
                                     worksheet, board[i * 7 + k])
                            setStyle(12 * gameID + i + 1, k +
                                     1, worksheet)  # (optional)

                    setValue(gameID * 12 + 8, 1, worksheet, "Steps:")
                    setValue(gameID * 12 + 8, 2, worksheet, move)
                    setValue(gameID * 12 + 9, 1, worksheet, "R Value:")
                    setValue(gameID * 12 + 10, 1, worksheet, "Y Value:")
                    setValue(1, 10, worksheet, gameID + 1)
                    workbook.save(file_path_xlsx)

                    # Erster Parameter bestimmt die Schriftart
                    # Zweiter Parameter bestimmt die Schriftgrösse
                    font = p.font.Font('freesansbold.ttf', 32)

                    if player:
                        # Text und koordinaten werden bestimmt
                        text = font.render(
                            'Gelb hat gewonnen!', True, (0, 255, 0), (0, 0, 255))
                        textRect = text.get_rect()
                        textRect.center = (180, height // 20)
                    else:
                        # Text und koordinaten werden bestimmt
                        text = font.render(
                            'Rot hat gewonnen!', True, (0, 255, 0), (0, 0, 255))
                        textRect = text.get_rect()
                        textRect.center = (160, height // 20)

                    # player wird auf None gesetzt, damit man nicht mehr setzen kann.
                    player = None
                    # Text wird gesetzt
                    screen.blit(text, textRect)

                    text = font.render('Neustarten', True,
                                       (0, 255, 0), (0, 0, 255))
                    textRect = text.get_rect()
                    textRect.center = (width - 110, height // 20)
                    screen.blit(text, textRect)

                if player is None:
                    position = p.mouse.get_pos()
                    # Es wird überprüft wo ob man auf "Neustarten" gedürckt hat
                    if textRect.x <= position[0] <= textRect.x + textRect.size[0]:
                        if textRect.y <= position[1] <= textRect.y + textRect.size[1]:
                            # Alle Werte werden zurückgesetzt, damit das Spiel von vorne beginnt
                            occupied = []
                            board = []
                            for n in range(42):
                                board.append("nnn")

                            move = 0
                            screen.fill(p.Color(0, 0, 0))
                            player = True

        if event.type == p.MOUSEBUTTONDOWN:
            if player is not None:
                # In played wird gespeichert ob jemand gesetzt hat.
                played = put(p.mouse.get_pos()[0] // sizeOne)
                # Spielzug wird um 1 erhöht
                move = move + 1
                # In status wird gespeichert ob jemand gewonnen hat.
                status = check_win(board)
            if played and player is not None:
                if not status:
                    if player:
                        # Spieler wird gewechselt
                        player = False
                    else:
                        # Spieler wird gewechselt
                        player = True

                # Jemand hat gewonnen
                else:
                    # Excel Datei abrufen
                    workbook = xl.load_workbook(file_path_xlsx)
                    worksheet = workbook['Stats']

                    # Aktuelle GameID abrufen
                    gameID = getValue(1, 10, worksheet)

                    # Alle Werte in Excel speichern
                    for i in range(6):
                        for k in range(7):
                            setValue(12 * gameID + i + 1, k + 1,
                                     worksheet, board[i * 7 + k])
                            setStyle(12 * gameID + i + 1, k +
                                     1, worksheet)  # (optional)

                    setValue(gameID * 12 + 8, 1, worksheet, "Steps:")
                    setValue(gameID * 12 + 8, 2, worksheet, move)
                    setValue(gameID * 12 + 9, 1, worksheet, "R Value:")
                    setValue(gameID * 12 + 10, 1, worksheet, "Y Value:")
                    setValue(1, 10, worksheet, gameID + 1)
                    workbook.save(file_path_xlsx)

                    # Erster Parameter bestimmt die Schriftart
                    # Zweiter Parameter bestimmt die Schriftgrösse
                    font = p.font.Font('freesansbold.ttf', 32)

                    if player:
                        # Text und koordinaten werden bestimmt
                        text = font.render(
                            'Gelb hat gewonnen!', True, (0, 255, 0), (0, 0, 255))
                        textRect = text.get_rect()
                        textRect.center = (180, height // 20)
                    else:
                        # Text und koordinaten werden bestimmt
                        text = font.render(
                            'Rot hat gewonnen!', True, (0, 255, 0), (0, 0, 255))
                        textRect = text.get_rect()
                        textRect.center = (160, height // 20)

                    # player wird auf None gesetzt, damit man nicht mehr setzen kann.
                    player = None
                    # Text wird gesetzt
                    screen.blit(text, textRect)

                    text = font.render('Neustarten', True,
                                       (0, 255, 0), (0, 0, 255))
                    textRect = text.get_rect()
                    textRect.center = (width - 110, height // 20)
                    screen.blit(text, textRect)

            if player is None:
                position = p.mouse.get_pos()
                # Es wird überprüft wo ob man auf "Neustarten" gedürckt hat
                if textRect.x <= position[0] <= textRect.x + textRect.size[0]:
                    if textRect.y <= position[1] <= textRect.y + textRect.size[1]:
                        # Alle Werte werden zurückgesetzt, damit das Spiel von vorne beginnt
                        occupied = []
                        board = []
                        for n in range(42):
                            board.append("nnn")
                        move = 0
                        screen.fill(p.Color(0, 0, 0))
                        player = True

        # Das Feld wird gezeichnet
        drawGame()

p.quit()
